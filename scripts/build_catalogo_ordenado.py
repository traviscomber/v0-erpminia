from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_DEFAULT = Path(r"C:\Users\juanf\Downloads\Base Existencias.xlsx")
OUTPUT_DEFAULT = Path("outputs") / "catalogo_base_existencias_ordenado.xlsx"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def slug_code(value: str, max_len: int = 6) -> str:
    text = strip_accents(normalize_text(value)).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return (text[:max_len] or "GEN").upper()


def normalize_family(value: str) -> str:
    key = strip_accents(normalize_text(value)).lower()
    mapping = {
        "repuesto": "Repuesto",
        "sondaje": "Sondaje",
        "ferreteria": "Ferretería",
        "filtro": "Filtro",
        "fitting": "Fitting",
        "acero": "Acero",
        "electrico": "Eléctrico",
        "compo": "Compo",
        "epp": "EPP",
        "rodamiento": "Rodamiento",
        "perforacion": "Perforación",
        "bomba": "Bomba",
        "correa": "Correa",
        "viveres": "Víveres",
        "neumatico": "Neumático",
        "soldadura": "Soldadura",
        "coraza": "Coraza",
        "reactivo": "Reactivo",
        "lubricante": "Lubricante",
        "escritorio": "Escritorio",
        "explosivo": "Explosivo",
        "combustible": "Combustible",
        "fortificacion": "Fortificación",
        "cinta": "Cinta",
        "malla": "Malla",
        "lona": "Lona",
        "madera": "Madera",
        "feria": "Feria",
        "laboratorio": "Laboratorio",
        "bola": "Bola",
    }
    return mapping.get(key, normalize_text(value) or "Sin clasificar")


FAMILY_CODE_MAP = {
    "Repuesto": "RPT",
    "Sondaje": "SON",
    "Ferretería": "FER",
    "Filtro": "FLT",
    "Fitting": "FIT",
    "Acero": "ACE",
    "Eléctrico": "ELC",
    "Compo": "CPO",
    "EPP": "EPP",
    "Rodamiento": "RDM",
    "Perforación": "PER",
    "Bomba": "BOM",
    "Correa": "CRR",
    "Víveres": "VIV",
    "Neumático": "NEU",
    "Soldadura": "SLD",
    "Coraza": "CRZ",
    "Reactivo": "RXT",
    "Lubricante": "LUB",
    "Escritorio": "EST",
    "Explosivo": "EXP",
    "Combustible": "COMB",
    "Fortificación": "FTF",
    "Cinta": "CNT",
    "Malla": "MLL",
    "Lona": "LON",
    "Madera": "MDR",
    "Feria": "FRA",
    "Laboratorio": "LAB",
    "Bola": "BOL",
    "Sin clasificar": "GEN",
}


MACRO_DEFS = [
    ("MNT", "Mantenimiento y repuestos", {"Repuesto", "Filtro", "Fitting", "Rodamiento", "Correa", "Coraza", "Bomba"}),
    ("MIN", "Operacion mina y perforacion", {"Sondaje", "Perforacion", "Explosivo", "Fortificacion", "Malla", "Reactivo", "Laboratorio"}),
    ("TAL", "Taller y metalmecanica", {"Ferreteria", "Acero", "Soldadura", "Compo", "Cinta", "Lona", "Madera", "Bola"}),
    ("ENE", "Energia y fluidos", {"Electrico", "Combustible", "Lubricante"}),
    ("ADM", "Administracion y consumo", {"EPP", "Viveres", "Escritorio", "Feria"}),
    ("TRA", "Transporte y rodados", {"Neumatico"}),
    ("QUI", "Quimicos y laboratorio", {"Reactivo", "Laboratorio"}),
    ("OBR", "Obras y construccion", {"Fortificacion", "Malla", "Acero", "Soldadura", "Lona"}),
]


def macro_for_family(family: str) -> tuple[str, str]:
    family_key = strip_accents(normalize_text(family)).lower()
    for code, label, families in MACRO_DEFS:
        normalized = {strip_accents(value).lower() for value in families}
        if family_key in normalized:
            return code, label
    return "OTR", "Otros"


def clean_route(route: object) -> str:
    value = normalize_text(route).replace("/", "\\").replace(">", "\\")
    value = re.sub(r"\\+", r"\\", value)
    return value.strip("\\")


def parse_date(value: object) -> str:
    if value in (None, "", "---"):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    text = normalize_text(value)
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            pass
    return text


def style_header(row):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )


def style_title(ws, title: str, subtitle: str):
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws["A2"].font = Font(size=10, color="6B7280")


def autosize(ws, max_col: int, max_width: int = 42):
    for col in range(1, max_col + 1):
        longest = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                longest = max(longest, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(longest + 2, 12), max_width)


def build_workbook(source_path: Path) -> Workbook:
    src = load_workbook(source_path, data_only=False)
    if "Productos" not in src.sheetnames or "Centros de Costos" not in src.sheetnames:
        raise ValueError("El XLS debe incluir las hojas 'Productos' y 'Centros de Costos'.")

    productos_ws = src["Productos"]
    centros_ws = src["Centros de Costos"]

    products = list(productos_ws.iter_rows(min_row=2, values_only=True))
    centers = list(centros_ws.iter_rows(min_row=2, values_only=True))

    product_rows = []
    family_counts = Counter()
    for row in products:
        codigo_origen = normalize_text(row[0])
        familia = normalize_family(row[1])
        subfamilia_src = normalize_text(row[2])
        equipo_src = normalize_text(row[3])
        producto = normalize_text(row[4])
        family_counts[familia] += 1
        product_rows.append(
            {
                "codigo_origen": codigo_origen,
                "familia": familia,
                "subfamilia_src": subfamilia_src,
                "equipo_src": equipo_src,
                "producto": producto,
            }
        )

    families = sorted(set(row["familia"] for row in product_rows), key=lambda fam: (-family_counts[fam], fam))

    family_meta = {}
    for family in families:
        macro_code, macro_label = macro_for_family(family)
        family_meta[family] = {
            "macro_code": macro_code,
            "macro_label": macro_label,
            "family_code": FAMILY_CODE_MAP.get(family, slug_code(family, 4)),
        }

    family_seq = defaultdict(int)
    normalized_products = []
    for row in sorted(product_rows, key=lambda r: (family_meta[r["familia"]]["macro_code"], r["familia"], r["producto"], r["codigo_origen"])):
        family_code = family_meta[row["familia"]]["family_code"]
        family_seq[row["familia"]] += 1
        standard_code = f"{family_code}-{family_seq[row['familia']]:04d}"
        normalized_products.append(
            {
                "COD_MACRO": family_meta[row["familia"]]["macro_code"],
                "MACRO_CATEGORIA": family_meta[row["familia"]]["macro_label"],
                "COD_FAMILIA": family_code,
                "FAMILIA": row["familia"],
                "CODIGO_ORIGEN": row["codigo_origen"],
                "CODIGO_ESTANDAR": standard_code,
                "PRODUCTO": row["producto"],
                "SUBFAMILIA": row["subfamilia_src"] or "Sin definir",
                "EQUIPO": row["equipo_src"] or "Sin definir",
                "ESTADO_CLASIFICACION": "Pendiente de subclasificación",
            }
        )

    center_rows = []
    root_counts = Counter()
    for row in centers:
        route = clean_route(row[8])
        segments = [seg for seg in route.split("\\") if seg]
        root = segments[0] if segments else ""
        root_counts[root] += 1
        center_rows.append(
            {
                "CREADOR_POR": normalize_text(row[0]),
                "CODIGO_ORIGEN": normalize_text(row[1]),
                "DISCONTINUADO": normalize_text(row[2]),
                "FECHA_CREACION": parse_date(row[3]),
                "FECHA_MODIFICACION": parse_date(row[4]),
                "MODIFICADO_POR": normalize_text(row[5]),
                "NOMBRE": normalize_text(row[6]),
                "NOTAS": normalize_text(row[7]),
                "RUTA_COMPLETA": route,
                "RAIZ": root,
                "NIVEL": len(segments),
                "CODIGO_PADRE": segments[-2] if len(segments) >= 2 else "",
                "ESTADO": "Inactivo" if normalize_text(row[2]).lower() == "si" else "Activo",
            }
        )

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Resumen")
    ws_catalog = wb.create_sheet("Catalogo_Codificado")
    ws_families = wb.create_sheet("Mapa_Familias")
    ws_centers = wb.create_sheet("Centros_Costos")

    style_title(ws_summary, "Base Existencias - Catálogo Codificado", "Versión limpia con códigos estándar, macros y trazabilidad.")
    summary_rows = [
        ["Métrica", "Valor"],
        ["Productos totales", len(normalized_products)],
        ["Familias normalizadas", len(families)],
        ["Centros de costo", len(center_rows)],
        ["Raíces de centros de costo", len(root_counts)],
        ["Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Fuente", str(source_path)],
    ]
    for r_idx, row in enumerate(summary_rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=value)
    style_header(ws_summary[4])
    ws_summary["A12"] = "Observaciones"
    ws_summary["A12"].font = Font(bold=True)
    ws_summary["A13"] = "SUBFAMILIA y EQUIPO venían vacíos en el XLS original; se dejan como 'Sin definir' para no inventar data."
    ws_summary["A13"].alignment = Alignment(wrap_text=True)
    ws_summary.freeze_panes = "A4"
    autosize(ws_summary, 2, 44)

    style_title(ws_catalog, "Catálogo de Productos Codificado", "Cada artículo conserva su código original y recibe un código estándar por familia.")
    catalog_headers = ["COD_MACRO", "MACRO_CATEGORIA", "COD_FAMILIA", "FAMILIA", "CODIGO_ORIGEN", "CODIGO_ESTANDAR", "PRODUCTO", "SUBFAMILIA", "EQUIPO", "ESTADO_CLASIFICACION"]
    for col_idx, header in enumerate(catalog_headers, start=1):
        ws_catalog.cell(row=4, column=col_idx, value=header)
    style_header(ws_catalog[4])
    for row_idx, row in enumerate(normalized_products, start=5):
        for col_idx, header in enumerate(catalog_headers, start=1):
            ws_catalog.cell(row=row_idx, column=col_idx, value=row[header])
    ws_catalog.freeze_panes = "A5"
    ws_catalog.auto_filter.ref = f"A4:{get_column_letter(len(catalog_headers))}{ws_catalog.max_row}"
    autosize(ws_catalog, len(catalog_headers), 40)

    style_title(ws_families, "Mapa de Familias", "Clasificación usada para ordenar el catálogo por macrogrupo y familia.")
    fam_headers = ["COD_MACRO", "MACRO_CATEGORIA", "COD_FAMILIA", "FAMILIA", "PRODUCTOS"]
    for col_idx, header in enumerate(fam_headers, start=1):
        ws_families.cell(row=4, column=col_idx, value=header)
    style_header(ws_families[4])
    for row_idx, family in enumerate(families, start=5):
        meta = family_meta[family]
        ws_families.cell(row=row_idx, column=1, value=meta["macro_code"])
        ws_families.cell(row=row_idx, column=2, value=meta["macro_label"])
        ws_families.cell(row=row_idx, column=3, value=meta["family_code"])
        ws_families.cell(row=row_idx, column=4, value=family)
        ws_families.cell(row=row_idx, column=5, value=family_counts[family])
    ws_families.freeze_panes = "A5"
    ws_families.auto_filter.ref = f"A4:E{ws_families.max_row}"
    autosize(ws_families, 5, 40)

    style_title(ws_centers, "Centros de Costo Normalizados", "Se conserva la estructura original del XLS con columnas limpias y trazabilidad de ruta.")
    center_headers = ["CREADOR_POR", "CODIGO_ORIGEN", "DISCONTINUADO", "FECHA_CREACION", "FECHA_MODIFICACION", "MODIFICADO_POR", "NOMBRE", "NOTAS", "RUTA_COMPLETA", "RAIZ", "NIVEL", "CODIGO_PADRE", "ESTADO"]
    for col_idx, header in enumerate(center_headers, start=1):
        ws_centers.cell(row=4, column=col_idx, value=header)
    style_header(ws_centers[4])
    for row_idx, row in enumerate(center_rows, start=5):
        for col_idx, header in enumerate(center_headers, start=1):
            ws_centers.cell(row=row_idx, column=col_idx, value=row[header])
    ws_centers.freeze_panes = "A5"
    ws_centers.auto_filter.ref = f"A4:M{ws_centers.max_row}"
    autosize(ws_centers, len(center_headers), 38)

    for ws in (ws_catalog, ws_families, ws_centers):
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin", color="E5E7EB"),
                    right=Side(style="thin", color="E5E7EB"),
                    top=Side(style="thin", color="E5E7EB"),
                    bottom=Side(style="thin", color="E5E7EB"),
                )
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    return wb


def main():
    parser = argparse.ArgumentParser(description="Build a cleaned and coded workbook from Base Existencias.xlsx")
    parser.add_argument("--input", default=str(SOURCE_DEFAULT))
    parser.add_argument("--output", default=str(OUTPUT_DEFAULT))
    args = parser.parse_args()

    source_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = build_workbook(source_path)
    workbook.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
